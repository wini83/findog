"""Sheet-level helpers to read, write and format payment data."""

import calendar
from copy import copy
from datetime import datetime, timedelta

from openpyxl.cell import Cell
from openpyxl.styles import Color, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from payment import Payment
from payment_category import PaymentCategory

RED_FILL = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
GREEN_FILL = PatternFill(fill_type='solid', start_color="92D050")
YELLOW_FILL = PatternFill(fill_type='solid', start_color=Color(indexed=5))
BLUE_FILL = PatternFill(fill_type='solid', start_color="6ED1FE")


def add_months(source_date: datetime, months: int):
    """Return a new date shifted by a number of months."""
    month = source_date.month - 1 + months
    year = source_date.year + month // 12
    month = month % 12 + 1
    day = min(source_date.day, calendar.monthrange(year, month)[1])
    return datetime(year, month, day)


class PaymentSheet:
    """Abstraction over an openpyxl Worksheet for payments."""

    _sheet: Worksheet = None
    _name: str = None
    _categories: dict[str, PaymentCategory] = None
    _monitored_cols: list[str] = None

    def __init__(self, worksheet: Worksheet, name: str, monitored_cols: list[str]):
        """Bind to a worksheet and configure monitored columns."""
        self._categories = {}
        self._sheet = worksheet
        self._name = name
        self._monitored_cols = monitored_cols

    @property
    def monitored_cols(self) -> list[str]:
        """Return the list of monitored column letters."""
        return self._monitored_cols

    @monitored_cols.setter
    def monitored_cols(self, value: list[str]):
        """Set the list of monitored column letters."""
        self._monitored_cols = value

    @property
    def categories(self) -> dict[str, PaymentCategory]:
        """Return a mapping from category name to PaymentCategory."""
        return self._categories

    @property
    def sheet(self) -> Worksheet:
        """Return the underlying worksheet."""
        return self._sheet

    @property
    def name(self):
        """Return the worksheet human-readable name."""
        return self._name

    @property
    def get_active_row(self) -> int:
        """Locate the current month row or return -1 if not found."""
        current_month = datetime.now().month
        current_year = datetime.now().year
        current_row = 2
        found: bool = False
        while self._sheet[f"A{current_row}"].value is not None:
            date_item = self._sheet[f"A{current_row}"].value
            if (date_item.month == current_month) and (date_item.year == current_year):
                found = True
                break
            current_row += 1
        if not found:
            current_row = -1
        return current_row

    def populate_categories(self, active_row: int):
        """Populate internal categories and payments from the active row up."""
        # NOTE: consider validation/sanitization if inputs can be external
        for column in self._monitored_cols:
            name = self._sheet[f"{column}1"].value
            item: PaymentCategory = PaymentCategory(name=name, column=column)
            comment = self._sheet[f"{column}1"].comment
            if comment is None:
                item.icon = "fa-camera"
            else:
                item.icon = comment.text.strip()
            processed_row = active_row
            while (
                self._sheet[f"{column}{processed_row}"].value is not None
                and processed_row > 1
            ):
                column_int, due_date, new_payment = self.populate_payment(
                    processed_row, column
                )
                item.payments[f'{due_date.year}-{due_date.month}'] = new_payment

                self._categories[item.name] = item
                self.format_payment(column_int, new_payment)
                processed_row -= 1
        self._format_this_month_cells(active_row)

    def populate_payment(self, active_row, column):
        """Read a single row/column triplet into a Payment instance."""
        amount = float(self._sheet[f"{column}{active_row}"].value)
        column_int = self._sheet[f"{column}{active_row}"].col_idx
        try:
            paid = bool(self._sheet.cell(column=column_int + 1, row=active_row).value)
        except ValueError:
            paid = False
        due_date: datetime = self._sheet.cell(
            column=column_int + 2, row=active_row
        ).value
        new_payment = Payment(
            paid=paid, due_date=due_date, amount=amount, excel_row=active_row
        )
        return column_int, due_date, new_payment

    # noinspection PyDunderSlots,PyUnresolvedReferences
    def _format_this_month_cells(self, active_row: int):
        """Highlight current and previous month summary cells."""
        cell_this_month_sum: Cell = self.sheet.cell(column=2, row=active_row)
        cell_this_month: Cell = self.sheet.cell(column=1, row=active_row)
        cell_this_month.fill = YELLOW_FILL
        cell_this_month_sum.fill = YELLOW_FILL
        # NOTE: consider adding checks for edge cases
        cell_previous_month_sum: Cell = self.sheet.cell(column=2, row=active_row - 1)
        cell_previous_month: Cell = self.sheet.cell(column=1, row=active_row - 1)
        cell_previous_month.fill = GREEN_FILL
        cell_previous_month_sum.fill = GREEN_FILL

    # noinspection PyDunderSlots,PyUnresolvedReferences
    def format_payment(self, column: int, payment: Payment):
        """Apply fills to amount/paid/due cells depending on status."""
        # NOTE: add input validation if needed
        cell_amount: Cell = self.sheet.cell(row=payment.excel_row, column=column)
        cell_paid: Cell = self.sheet.cell(row=payment.excel_row, column=column + 1)
        cell_due_date: Cell = self.sheet.cell(row=payment.excel_row, column=column + 2)
        if payment.overdue:
            cell_amount.fill = RED_FILL
            cell_paid.fill = RED_FILL
            cell_due_date.fill = RED_FILL
        elif payment.paid:
            cell_amount.fill = GREEN_FILL
            cell_paid.fill = GREEN_FILL
            cell_due_date.fill = GREEN_FILL
        elif (
            payment.due_date.year == datetime.now().year
            and payment.due_date.month == datetime.now().month
        ):
            cell_amount.fill = YELLOW_FILL
            cell_paid.fill = YELLOW_FILL
            cell_due_date.fill = YELLOW_FILL
        else:
            cell_amount.fill = BLUE_FILL
            cell_paid.fill = BLUE_FILL
            cell_due_date.fill = BLUE_FILL

    def populate_next_month(self, current_row: int):
        """Ensure next month row is present and pre-filled based on current."""
        proceed_further = self._process_next_month_cell(current_row)
        if proceed_further:
            self._process_next_sum(current_row)
            self._populate_next_month_payments(current_row)

    # noinspection PyDunderSlots,PyUnresolvedReferences
    def _populate_next_month_payments(self, current_row: int):
        """Copy amount/paid/due values forward into next month row."""
        for category in self.categories.values():
            current_amount_cell: Cell = self.sheet[f'{category.column}{current_row}']
            current_paid_cell = self.sheet.cell(
                row=current_row, column=current_amount_cell.col_idx + 1
            )
            current_due_date_cell = self.sheet.cell(
                row=current_row, column=current_amount_cell.col_idx + 2
            )

            next_amount_cell: Cell = self.sheet[f'{category.column}{current_row + 1}']
            next_paid_cell = self.sheet.cell(
                row=current_row + 1, column=current_amount_cell.col_idx + 1
            )
            next_due_date_cell = self.sheet.cell(
                row=current_row + 1, column=current_amount_cell.col_idx + 2
            )
            if next_amount_cell.value is None:
                next_amount_cell.value = current_amount_cell.value
                next_amount_cell.number_format = current_amount_cell.number_format
                next_amount_cell.font = copy(current_amount_cell.font)
                next_amount_cell.border = copy(current_amount_cell.border)
            next_amount_cell.fill = BLUE_FILL

            if next_paid_cell.value is None:
                next_paid_cell.value = 0
                next_paid_cell.number_format = current_paid_cell.number_format
                next_paid_cell.font = copy(current_paid_cell.font)
                next_paid_cell.border = copy(current_paid_cell.border)
            next_paid_cell.fill = BLUE_FILL

            if next_due_date_cell.value is None:
                next_due_date_cell.value = add_months(current_due_date_cell.value, 1)
                next_due_date_cell.number_format = current_due_date_cell.number_format
                next_due_date_cell.font = copy(current_due_date_cell.font)
                next_due_date_cell.border = copy(current_due_date_cell.border)
            next_due_date_cell.fill = BLUE_FILL

    def _generate_sum_string(self, row: int):
        """Build an Excel SUM formula over monitored columns for the row."""
        result = "=SUM("
        counter: int = 0
        for column_str in self._monitored_cols:
            if counter > 0:
                result += ","
            result += f'{column_str}{row}'
            counter += 1
        result += ")"
        return result

    # noinspection PyDunderSlots,PyUnresolvedReferences
    def _process_next_sum(self, current_row):
        """Create next month's summary formula if absent."""
        cell_this_month_sum: Cell = self.sheet.cell(column=2, row=current_row)
        cell_next_month_sum: Cell = self.sheet.cell(column=2, row=current_row + 1)
        if cell_next_month_sum.value is None:
            cell_next_month_sum.value = self._generate_sum_string(current_row + 1)
            cell_next_month_sum.number_format = cell_this_month_sum.number_format
            cell_next_month_sum.font = copy(cell_this_month_sum.font)
            cell_next_month_sum.border = copy(cell_this_month_sum.border)
            cell_next_month_sum.fill = BLUE_FILL

    # noinspection PyDunderSlots,PyUnresolvedReferences
    def _process_next_month_cell(self, current_row):
        """Create the next month date cell if missing; return whether to proceed."""
        cell_next_month: Cell = self.sheet.cell(column=1, row=current_row + 1)
        cell_this_month: Cell = self.sheet.cell(column=1, row=current_row)
        now = datetime.now()
        now = datetime(year=now.year, month=now.month, day=now.day)
        next_month_date = (now.replace(day=1) + timedelta(days=32)).replace(day=1)
        if cell_next_month.value is None:
            cell_next_month.value = next_month_date
            cell_next_month.number_format = self.sheet.cell(
                column=1, row=current_row
            ).number_format
            cell_next_month.number_format = cell_this_month.number_format
            cell_next_month.font = copy(cell_this_month.font)
            cell_next_month.border = copy(cell_this_month.border)
            cell_next_month.fill = BLUE_FILL
            process = True
        elif cell_next_month.value == next_month_date:
            cell_next_month.fill = BLUE_FILL
            process = True
        else:
            process = False
        return process
