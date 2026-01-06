"""Workbook access and operations for payments tracking."""

from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell

from payment_list_item import PaymentListItem
from payment_sheet import PaymentSheet


def sort_payment_list_by_date(
    item_list: list[PaymentListItem],
) -> list[PaymentListItem]:
    """Return items sorted by payment due date ascending."""
    return sorted(item_list, key=lambda x: x.payment.due_date, reverse=False)


def make_json_payments(item_list: list[PaymentListItem]):
    """Serialize list items into JSON-friendly list of dicts."""
    result = []
    for item3 in item_list:
        result.append(item3.to_dict())
    return result


class PaymentBook:
    """High-level API to read/update monitored payment sheets in a workbook."""

    _payment_sheets: dict[str, PaymentSheet] = None
    _workbook: Workbook = None
    _monitored_sheets = None

    def __init__(self, monitored_sheets):
        """Initialize with a map of monitored sheet names to column letters."""
        self._payment_sheets = {}
        self._monitored_sheets = monitored_sheets

    def load_and_process(self, file: bytes):
        """Load workbook from bytes and build internal structures."""
        self._workbook = load_workbook(filename=BytesIO(file), keep_vba=True)
        for sheet_name, monit_cats in self.monitored_sheets.items():
            if sheet_name in self._workbook:
                payment_sheet = PaymentSheet(
                    self._workbook[sheet_name], sheet_name, monit_cats
                )
                current_row = payment_sheet.get_active_row
                if current_row > 1:
                    payment_sheet.populate_categories(current_row)
                    payment_sheet.populate_next_month(current_row)
                self._payment_sheets[sheet_name] = payment_sheet

    def save_to_file(self, filename):
        """Persist current workbook to a file path."""
        self._workbook.save(filename=filename)

    def update_current_payment(
        self,
        sheet_name: str,
        category_name: str,
        amount: float = None,
        paid: bool = None,
        due_date: datetime = None,
        force_unpaid: bool = None,
    ):
        """Update fields in the current month row for a given category."""
        # NOTE: this method can be refactored further if API allows
        # pylint: disable=too-many-arguments,too-many-positional-arguments
        sheet: PaymentSheet = self._payment_sheets.get(sheet_name)
        if not sheet or sheet.name != sheet_name:
            return

        cat = sheet.categories.get(category_name)
        if not cat or cat.name != category_name:
            return

        now = datetime.now()
        pmt = cat.payments.get(f'{now.year}-{now.month}')
        if not pmt:
            return

        if not (
            pmt.due_date.year == date.today().year
            and pmt.due_date.month == date.today().month
        ):
            return

        cell_amount: Cell = sheet.sheet[f'{cat.column}{pmt.excel_row}']
        cell_paid: Cell = sheet.sheet.cell(
            row=cell_amount.row, column=cell_amount.column + 1
        )
        cell_due_date: Cell = sheet.sheet.cell(
            row=cell_amount.row, column=cell_amount.column + 2
        )

        if force_unpaid is None:
            force_unpaid = True

        if amount is not None:
            pmt.amount = amount
            cell_amount.value = amount

        if paid is not None:
            if not paid:
                if force_unpaid:
                    cell_paid.value = int(paid)
                    pmt.paid = paid
            else:
                pmt.paid = paid
                cell_paid.value = int(paid)

        if due_date is not None:
            pmt.due_date = due_date
            cell_due_date.value = due_date
            sheet.format_payment(cell_amount.column, pmt)

    @property
    def payment_list(self) -> list[PaymentListItem]:
        """Flatten all sheet categories into a single list of items."""
        result: list[PaymentListItem] = []
        for sheet in self._payment_sheets.values():
            for category in sheet.categories.values():
                for payment in category.payments.values():
                    new_item: PaymentListItem = PaymentListItem(
                        payment, category, sheet
                    )
                    result.append(new_item)
        return result

    @property
    def sheets(self):
        """Access the internal mapping of sheet name to PaymentSheet."""
        return self._payment_sheets

    @property
    def monitored_sheets(self):
        """Get monitored sheet configuration."""
        return self._monitored_sheets

    @monitored_sheets.setter
    def monitored_sheets(self, value):
        """Set monitored sheet configuration."""
        self._monitored_sheets = value
