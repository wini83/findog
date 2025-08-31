from datetime import datetime
from io import BytesIO

from openpyxl import Workbook

from payment_book import PaymentBook


def _make_workbook_bytes():
    wb = Workbook()
    ws = wb.active
    ws.title = "Home"

    # Headers
    ws["A1"].value = "Date"
    ws["B1"].value = "Sum"
    ws["C1"].value = "Rent"

    # Current month row
    now = datetime.now()
    ws["A2"].value = datetime(year=now.year, month=now.month, day=1)
    ws["B2"].value = 0
    ws["C2"].value = 1500.0
    ws["D2"].value = 0
    ws["E2"].value = now

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


class TestPaymentBook:
    def test_load_and_update(self):
        data = _make_workbook_bytes()
        book = PaymentBook(monitored_sheets={"Home": ["C"]})
        book.load_and_process(data)

        # Payment structures populated
        assert "Home" in book.sheets
        items = book.payment_list
        assert len(items) == 1

        # Update current payment: mark as paid and change amount/due date
        new_amount = 1600.0
        new_due = datetime.now()
        book.update_current_payment(
            sheet_name="Home",
            category_name="Rent",
            amount=new_amount,
            paid=True,
            due_date=new_due,
        )

        # Validate updated state via list
        items = book.payment_list
        assert len(items) == 1
        item = items[0]
        assert item.payment.amount == new_amount
        assert item.payment.paid is True
        assert item.payment.due_date.year == new_due.year
