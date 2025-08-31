from datetime import datetime

from openpyxl import Workbook

from payment_sheet import PaymentSheet


def _setup_basic_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "Home"

    # Header row
    ws["A1"].value = "Date"
    ws["B1"].value = "Sum"
    ws["C1"].value = "Rent"

    # Current month row (row 2)
    now = datetime.now()
    ws["A2"].value = datetime(year=now.year, month=now.month, day=1)
    ws["B2"].value = 0
    ws["C2"].value = 1200.0  # amount
    ws["D2"].value = 0  # paid
    ws["E2"].value = now  # due_date

    return wb, ws


class TestPaymentSheet:
    def test_populate_and_get_active_row(self):
        wb, ws = _setup_basic_sheet()
        sheet = PaymentSheet(ws, ws.title, ["C"])  # monitor column C

        active_row = sheet.get_active_row
        assert active_row == 2

        sheet.populate_categories(active_row)
        assert "Rent" in sheet.categories
        cat = sheet.categories["Rent"]
        now = datetime.now()
        key = f"{now.year}-{now.month}"
        assert key in cat.payments
        pmt = cat.payments[key]
        assert pmt.amount == 1200.0
        assert not pmt.paid
        assert pmt.due_date.year == now.year and pmt.due_date.month == now.month

    def test_populate_next_month(self):
        wb, ws = _setup_basic_sheet()
        sheet = PaymentSheet(ws, ws.title, ["C"])  # monitor column C

        active_row = sheet.get_active_row
        sheet.populate_categories(active_row)
        sheet.populate_next_month(active_row)

        # Next row should be created
        next_row = active_row + 1
        assert ws.cell(column=1, row=next_row).value is not None  # A3 date exists
        # Sum formula should be present in B3
        assert ws.cell(column=2, row=next_row).value.startswith("=SUM(")
        # Amount/paid/due date carried forward
        assert ws["C3"].value == ws["C2"].value
        assert ws["D3"].value == 0
        assert ws["E3"].value is not None
